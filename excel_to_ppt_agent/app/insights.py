from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


@dataclass
class ChartSpec:
    chart_type: str  # 'line', 'bar', 'clustered_bar'
    title: str
    category_labels: List[Any]
    series: Dict[str, List[float]]  # series name -> values
    x_axis_title: Optional[str] = None
    y_axis_title: Optional[str] = None


@dataclass
class AnalysisResult:
    source_path: str
    rows: List[Dict[str, Any]]
    headers: List[str]
    date_column: Optional[str]
    numeric_columns: List[str]
    categorical_columns: List[str]
    summary_stats: Dict[str, Dict[str, float]]
    key_insights: List[str]
    charts: List[ChartSpec] = field(default_factory=list)


def _read_excel_rows(excel_path: str | Path, sheet: Optional[str | int]) -> Tuple[List[str], List[Dict[str, Any]]]:
    wb = load_workbook(filename=str(excel_path), data_only=True, read_only=True)
    if sheet is None:
        ws = wb.worksheets[0]
    else:
        ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return [], []
    rows: List[Dict[str, Any]] = []
    for r in rows_iter:
        row_dict = {headers[i]: r[i] for i in range(len(headers))}
        # Skip fully empty rows
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row_dict.values()):
            continue
        rows.append(row_dict)
    return headers, rows


def _is_parseable_float(value: Any) -> bool:
    try:
        float(value)
        return True
    except Exception:
        return False


def _infer_date_column(headers: List[str], rows: List[Dict[str, Any]]) -> Optional[str]:
    candidate_names = {"date", "datetime", "period", "time"}
    for h in headers:
        if h and h.strip().lower() in candidate_names:
            return h
    # fallback: first column where >=50% of non-null values parse as datetime
    def parse_dt(x: Any) -> Optional[datetime]:
        if isinstance(x, datetime):
            return x
        if isinstance(x, str):
            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
                try:
                    return datetime.strptime(x.strip(), fmt)
                except Exception:
                    continue
        return None

    for h in headers:
        values = [row.get(h) for row in rows]
        non_null = [v for v in values if v is not None and str(v).strip() != ""]
        if not non_null:
            continue
        hits = sum(1 for v in non_null if parse_dt(v) is not None)
        if hits / max(1, len(non_null)) >= 0.5:
            return h
    return None


def _select_columns(headers: List[str], rows: List[Dict[str, Any]]) -> tuple[List[str], List[str]]:
    total = max(1, len(rows))
    numeric_columns: List[str] = []
    categorical_columns: List[str] = []
    for h in headers:
        values = [row.get(h) for row in rows]
        non_null = [v for v in values if v is not None and str(v).strip() != ""]
        if not non_null:
            continue
        parse_ratio = sum(1 for v in non_null if _is_parseable_float(v)) / len(non_null)
        if parse_ratio >= 0.8:
            numeric_columns.append(h)
        else:
            unique = len(set(map(str, non_null)))
            if unique <= max(50, int(0.1 * total)):
                categorical_columns.append(h)
    return numeric_columns, categorical_columns


def _to_float(value: Any) -> float:
    try:
        return float(value)
    except Exception:
        return 0.0


def _parse_month_key(v: Any) -> Optional[str]:
    if isinstance(v, datetime):
        return v.strftime("%Y-%m")
    if isinstance(v, str) and v.strip():
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(v.strip(), fmt).strftime("%Y-%m")
            except Exception:
                continue
    return None


def _build_key_insights(
    headers: List[str], rows: List[Dict[str, Any]], date_column: Optional[str], numeric_columns: List[str], categorical_columns: List[str]
) -> List[str]:
    insights: List[str] = []
    if not numeric_columns:
        return ["No numeric columns detected. Provide a sheet with numeric metrics (e.g., Sales, Units)."]

    metric = numeric_columns[0]
    total_val = sum(_to_float(r.get(metric)) for r in rows)
    insights.append(f"Total {metric}: {total_val:,.0f}")

    if date_column is not None:
        monthly: Dict[str, float] = {}
        for r in rows:
            mk = _parse_month_key(r.get(date_column))
            if mk is None:
                continue
            monthly[mk] = monthly.get(mk, 0.0) + _to_float(r.get(metric))
        if len(monthly) >= 2:
            ordered = [k for k in sorted(monthly.keys())]
            first, last = monthly[ordered[0]], monthly[ordered[-1]]
            base = first if first != 0 else 1.0
            change = (last - first) / base
            direction = "up" if change >= 0 else "down"
            insights.append(f"Trend: {direction} {abs(change)*100:.1f}% from first to last month")

    if categorical_columns:
        cat = categorical_columns[0]
        by_cat: Dict[str, float] = {}
        for r in rows:
            key = str(r.get(cat)) if r.get(cat) is not None else ""
            by_cat[key] = by_cat.get(key, 0.0) + _to_float(r.get(metric))
        top3 = sorted(by_cat.items(), key=lambda kv: kv[1], reverse=True)[:3]
        parts = ", ".join(f"{k}: {v:,.0f}" for k, v in top3)
        insights.append(f"Top {cat} by {metric}: {parts}")

    return insights


def _charts(
    headers: List[str], rows: List[Dict[str, Any]], date_column: Optional[str], numeric_columns: List[str], categorical_columns: List[str]
) -> List[ChartSpec]:
    charts: List[ChartSpec] = []
    if not numeric_columns:
        return charts
    metric = numeric_columns[0]

    if date_column is not None:
        monthly: Dict[str, float] = {}
        for r in rows:
            mk = _parse_month_key(r.get(date_column))
            if mk is None:
                continue
            monthly[mk] = monthly.get(mk, 0.0) + _to_float(r.get(metric))
        if monthly:
            labels = sorted(monthly.keys())
            charts.append(
                ChartSpec(
                    chart_type="line",
                    title=f"Monthly {metric}",
                    category_labels=labels,
                    series={metric: [monthly[l] for l in labels]},
                    x_axis_title="Month",
                    y_axis_title=metric,
                )
            )

    if categorical_columns:
        cat = categorical_columns[0]
        by_cat: Dict[str, float] = {}
        for r in rows:
            key = str(r.get(cat)) if r.get(cat) is not None else ""
            by_cat[key] = by_cat.get(key, 0.0) + _to_float(r.get(metric))
        if by_cat:
            top_items = sorted(by_cat.items(), key=lambda kv: kv[1], reverse=True)[:10]
            charts.append(
                ChartSpec(
                    chart_type="bar",
                    title=f"{metric} by {cat}",
                    category_labels=[k for k, _ in top_items],
                    series={metric: [v for _, v in top_items]},
                    x_axis_title=cat,
                    y_axis_title=metric,
                )
            )

    if len(categorical_columns) >= 2:
        cat_a, cat_b = categorical_columns[:2]
        # Build pivot
        pivot: Dict[str, Dict[str, float]] = {}
        for r in rows:
            a = str(r.get(cat_a)) if r.get(cat_a) is not None else ""
            b = str(r.get(cat_b)) if r.get(cat_b) is not None else ""
            pivot.setdefault(a, {})
            pivot[a][b] = pivot[a].get(b, 0.0) + _to_float(r.get(metric))
        if pivot:
            a_labels = sorted(pivot.keys())
            b_labels = sorted({b for inner in pivot.values() for b in inner.keys()})
            if len(b_labels) <= 8:
                series = {b: [pivot.get(a, {}).get(b, 0.0) for a in a_labels] for b in b_labels}
                charts.append(
                    ChartSpec(
                        chart_type="clustered_bar",
                        title=f"{metric} by {cat_a} and {cat_b}",
                        category_labels=a_labels,
                        series=series,
                        x_axis_title=cat_a,
                        y_axis_title=metric,
                    )
                )

    return charts


def _summary_stats(numeric_columns: List[str], rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, float]]:
    stats: Dict[str, Dict[str, float]] = {}
    for col in numeric_columns:
        values = [_to_float(r.get(col)) for r in rows if r.get(col) is not None and str(r.get(col)).strip() != ""]
        if not values:
            continue
        n = len(values)
        s = sum(values)
        mean = s / n
        # simple std (population)
        var = sum((v - mean) ** 2 for v in values) / n
        stats[col] = {
            "count": float(n),
            "mean": mean,
            "std": var ** 0.5,
            "min": min(values),
            "max": max(values),
            "sum": s,
        }
    return stats


def analyze_excel(excel_path: str | Path, sheet: Optional[str | int] = None) -> AnalysisResult:
    excel_path = str(excel_path)
    headers, rows = _read_excel_rows(excel_path, sheet)
    date_column = _infer_date_column(headers, rows) if headers else None
    numeric_columns, categorical_columns = _select_columns(headers, rows)
    key_insights = _build_key_insights(headers, rows, date_column, numeric_columns, categorical_columns)
    charts = _charts(headers, rows, date_column, numeric_columns, categorical_columns)
    stats = _summary_stats(numeric_columns, rows)
    return AnalysisResult(
        source_path=excel_path,
        rows=rows,
        headers=headers,
        date_column=date_column,
        numeric_columns=numeric_columns,
        categorical_columns=categorical_columns,
        summary_stats=stats,
        key_insights=key_insights,
        charts=charts,
    )

